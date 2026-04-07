#!/usr/bin/env python3
"""
Impala Table Health Checker
- Reads db.tablename list from a txt file
- Runs SELECT queries in parallel against 4 Impala clusters
- Outputs results as a formatted table and Excel file
"""

import sys
import re
import time
import argparse
import concurrent.futures
from datetime import datetime
from pathlib import Path

try:
    from impala.dbapi import connect
    IMPALA_AVAILABLE = True
except ImportError:
    IMPALA_AVAILABLE = False
    print("[WARN] impyla not installed. Running in DEMO mode.")

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARN] openpyxl not installed. Excel export disabled.")


# ── Configuration ─────────────────────────────────────────────────────────────

CLUSTERS = [
    {"name": "Cluster-A", "host": "impala-a.example.com", "port": 21050},
    {"name": "Cluster-B", "host": "impala-b.example.com", "port": 21050},
    {"name": "Cluster-C", "host": "impala-c.example.com", "port": 21050},
    {"name": "Cluster-D", "host": "impala-d.example.com", "port": 21050},
]

CONNECT_TIMEOUT = 30   # seconds
QUERY_TIMEOUT   = 60   # seconds

# Key error patterns to extract (checked in order, first match wins)
ERROR_PATTERNS = [
    (r"BlockMissingException",          "BlockMissingException"),
    (r"TableLoadingException",          "TableLoadingException"),
    (r"TableNotFoundException",         "TableNotFoundException"),
    (r"AnalysisException",              "AnalysisException"),
    (r"AuthorizationException",         "AuthorizationException"),
    (r"ImpalaRuntimeException",         "ImpalaRuntimeException"),
    (r"DatabaseNotFoundException",      "DatabaseNotFoundException"),
    (r"FileNotFoundException",          "FileNotFoundException"),
    (r"HdfsIOException",                "HdfsIOException"),
    (r"connection\s+refused",           "ConnectionRefused"),
    (r"timed?\s*out",                   "Timeout"),
    (r"permission\s+denied",            "PermissionDenied"),
]


# ── Error extraction ──────────────────────────────────────────────────────────

def extract_error_key(error_msg: str) -> str:
    """Return the most relevant short error label from an exception message."""
    for pattern, label in ERROR_PATTERNS:
        if re.search(pattern, error_msg, re.IGNORECASE):
            return label
    # Fallback: first non-empty line, truncated
    first_line = next((l.strip() for l in error_msg.splitlines() if l.strip()), error_msg)
    return first_line[:60]


# ── Query execution ───────────────────────────────────────────────────────────

def check_table(cluster: dict, table: str) -> dict:
    """
    Execute `SELECT 1 FROM <table> LIMIT 1` against one cluster.
    Returns {"cluster": ..., "table": ..., "status": "PASS"|"FAIL", "error": ...}
    """
    result = {"cluster": cluster["name"], "table": table, "status": None, "error": None}

    if not IMPALA_AVAILABLE:
        # Demo mode: simulate results
        import random
        time.sleep(random.uniform(0.1, 0.4))
        if random.random() < 0.25:
            fake_errors = [
                "BlockMissingException: Could not obtain block",
                "TableLoadingException: Failed to load metadata",
                "AnalysisException: Could not resolve table reference",
                "connection refused",
            ]
            err = random.choice(fake_errors)
            result["status"] = "FAIL"
            result["error"]  = extract_error_key(err)
        else:
            result["status"] = "PASS"
        return result

    try:
        conn = connect(
            host=cluster["host"],
            port=cluster["port"],
            timeout=CONNECT_TIMEOUT,
        )
        cur = conn.cursor()
        cur.execute(f"SELECT 1 FROM {table} LIMIT 1")
        cur.fetchone()
        cur.close()
        conn.close()
        result["status"] = "PASS"
    except Exception as e:
        result["status"] = "FAIL"
        result["error"]  = extract_error_key(str(e))
    return result


# ── Parallel execution ────────────────────────────────────────────────────────

def run_checks(clusters: list, tables: list, max_workers: int = 16) -> list:
    """Run all cluster×table checks in parallel, return flat list of results."""
    jobs = [(c, t) for c in clusters for t in tables]
    results = []
    total = len(jobs)
    done  = 0

    print(f"\nChecking {len(tables)} table(s) × {len(clusters)} cluster(s) "
          f"= {total} queries ...\n")

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(check_table, c, t): (c, t) for c, t in jobs}
        for future in concurrent.futures.as_completed(future_map):
            res = future.result()
            results.append(res)
            done += 1
            status_icon = "✓" if res["status"] == "PASS" else "✗"
            err_part    = f"  [{res['error']}]" if res["error"] else ""
            print(f"  [{done:>4}/{total}] {status_icon} {res['cluster']:<12} {res['table']}{err_part}")

    return results


# ── Build result matrix ───────────────────────────────────────────────────────

def build_matrix(clusters: list, tables: list, results: list) -> dict:
    """
    Returns:
      matrix[cluster_name][table] = {"status": ..., "error": ...}
    """
    matrix = {c["name"]: {} for c in clusters}
    for r in results:
        matrix[r["cluster"]][r["table"]] = {
            "status": r["status"],
            "error":  r["error"],
        }
    return matrix


# ── Console table output ──────────────────────────────────────────────────────

def print_table(clusters: list, tables: list, matrix: dict):
    COL_W = 22
    NAME_W = 14

    cluster_names = [c["name"] for c in clusters]

    # Header
    header = f"{'':>{NAME_W}} " + "  ".join(t[:COL_W].ljust(COL_W) for t in tables) + "  PASS/TOTAL"
    print("\n" + "=" * len(header))
    print(header)
    print("=" * len(header))

    col_pass  = {t: 0 for t in tables}
    col_total = {t: 0 for t in tables}

    for cname in cluster_names:
        row_pass = row_total = 0
        cells = []
        for t in tables:
            info = matrix[cname].get(t, {"status": "N/A", "error": None})
            st   = info["status"]
            err  = info["error"] or ""
            col_total[t] += 1
            row_total    += 1
            if st == "PASS":
                col_pass[t] += 1
                row_pass    += 1
                cell = "PASS"
            elif st == "FAIL":
                cell = f"FAIL({err})"[:COL_W]
            else:
                cell = st
            cells.append(cell.ljust(COL_W))
        row_summary = f"{row_pass}/{row_total}"
        print(f"{cname:>{NAME_W}} " + "  ".join(cells) + f"  {row_summary}")

    # Column summary
    print("-" * len(header))
    summary_cells = [f"{col_pass[t]}/{col_total[t]}".ljust(COL_W) for t in tables]
    total_pass  = sum(col_pass.values())
    total_total = sum(col_total.values())
    print(f"{'PASS/TOTAL':>{NAME_W}} " + "  ".join(summary_cells) + f"  {total_pass}/{total_total}")
    print("=" * len(header))
    print(f"\nOverall: {total_pass}/{total_total} PASS  |  "
          f"{total_total - total_pass}/{total_total} FAIL\n")


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(clusters: list, tables: list, matrix: dict, out_path: str):
    if not EXCEL_AVAILABLE:
        print("[WARN] openpyxl not available — skipping Excel export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table Check"

    # Colour palette
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    BLUE   = PatternFill("solid", fgColor="BDD7EE")
    GREY   = PatternFill("solid", fgColor="D9D9D9")
    WHITE  = PatternFill("solid", fgColor="FFFFFF")

    bold      = Font(bold=True)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    cluster_names = [c["name"] for c in clusters]

    # ── Row 1: header (cluster \ table | table1 | table2 | … | PASS/TOTAL) ──
    ws.cell(1, 1, "Cluster \\ Table").font = bold
    ws.cell(1, 1).fill      = GREY
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).border    = border

    for ci, t in enumerate(tables, start=2):
        cell = ws.cell(1, ci, t)
        cell.font      = bold
        cell.fill      = BLUE
        cell.alignment = center
        cell.border    = border

    summary_col = len(tables) + 2
    cell = ws.cell(1, summary_col, "PASS / TOTAL")
    cell.font      = bold
    cell.fill      = GREY
    cell.alignment = center
    cell.border    = border

    # ── Data rows ──────────────────────────────────────────────────────────
    col_pass  = {t: 0 for t in tables}
    col_total = {t: 0 for t in tables}

    for ri, cname in enumerate(cluster_names, start=2):
        row_pass = row_total = 0

        cell = ws.cell(ri, 1, cname)
        cell.font      = bold
        cell.fill      = BLUE
        cell.alignment = center
        cell.border    = border

        for ci, t in enumerate(tables, start=2):
            info = matrix[cname].get(t, {"status": "N/A", "error": None})
            st   = info["status"]
            err  = info["error"] or ""
            col_total[t] += 1
            row_total    += 1

            if st == "PASS":
                col_pass[t] += 1
                row_pass    += 1
                label = "PASS"
                fill  = GREEN
            elif st == "FAIL":
                label = f"FAIL\n{err}" if err else "FAIL"
                fill  = RED
            else:
                label = st
                fill  = WHITE

            cell = ws.cell(ri, ci, label)
            cell.fill      = fill
            cell.alignment = center
            cell.border    = border

        # Row summary
        cell = ws.cell(ri, summary_col, f"{row_pass}/{row_total}")
        cell.font      = bold
        cell.fill      = GREEN if row_pass == row_total else (RED if row_pass == 0 else YELLOW)
        cell.alignment = center
        cell.border    = border

    # ── Column summary row ──────────────────────────────────────────────────
    summary_row = len(cluster_names) + 2
    cell = ws.cell(summary_row, 1, "PASS / TOTAL")
    cell.font      = bold
    cell.fill      = GREY
    cell.alignment = center
    cell.border    = border

    total_pass = total_total = 0
    for ci, t in enumerate(tables, start=2):
        p = col_pass[t]
        n = col_total[t]
        total_pass  += p
        total_total += n
        cell = ws.cell(summary_row, ci, f"{p}/{n}")
        cell.font      = bold
        cell.fill      = GREEN if p == n else (RED if p == 0 else YELLOW)
        cell.alignment = center
        cell.border    = border

    # Grand total cell
    cell = ws.cell(summary_row, summary_col, f"{total_pass}/{total_total}")
    cell.font      = Font(bold=True, size=12)
    cell.fill      = GREEN if total_pass == total_total else (RED if total_pass == 0 else YELLOW)
    cell.alignment = center
    cell.border    = border

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 16
    for ci in range(2, len(tables) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 26
    ws.column_dimensions[openpyxl.utils.get_column_letter(summary_col)].width = 14

    # Freeze header row + cluster column
    ws.freeze_panes = "B2"

    wb.save(out_path)
    print(f"[Excel] Saved → {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Impala table health checker")
    parser.add_argument("table_file", nargs="?", default="tables.txt",
                        help="Path to txt file with db.tablename per line (default: tables.txt)")
    parser.add_argument("--workers", type=int, default=16,
                        help="Max parallel threads (default: 16)")
    parser.add_argument("--output", default="",
                        help="Excel output path (default: result_YYYYMMDD_HHMMSS.xlsx)")
    args = parser.parse_args()

    # Read table list
    table_file = Path(args.table_file)
    if not table_file.exists():
        print(f"[ERROR] Table file not found: {table_file}")
        sys.exit(1)

    tables = []
    with open(table_file, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                tables.append(line)

    if not tables:
        print("[ERROR] No tables found in the file.")
        sys.exit(1)

    print(f"Loaded {len(tables)} table(s) from {table_file}")

    # Run checks
    start   = time.time()
    results = run_checks(CLUSTERS, tables, max_workers=args.workers)
    elapsed = time.time() - start

    # Build matrix and print
    matrix = build_matrix(CLUSTERS, tables, results)
    print_table(CLUSTERS, tables, matrix)
    print(f"Elapsed: {elapsed:.1f}s")

    # Excel export
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = args.output or f"result_{ts}.xlsx"
    export_excel(CLUSTERS, tables, matrix, out_path)


if __name__ == "__main__":
    main()
